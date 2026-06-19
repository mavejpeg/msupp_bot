from __future__ import annotations

from datetime import date
from pathlib import Path
from tempfile import NamedTemporaryFile

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import Transaction
from app.services.budget import month_summary
from app.utils import money


async def export_xlsx(session: AsyncSession, month: date) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    s = await month_summary(session, month)

    ws.append(["Семейный бюджет", s["month"].strftime("%Y-%m")])
    ws.append([])
    ws.append(["План дохода", float(s["planned_income"])])
    ws.append(["Факт дохода", float(s["income"])])
    ws.append(["План расходов", float(s["planned_expense"])])
    ws.append(["Факт расходов", float(s["expense"])])
    ws.append(["Остаток бюджета", float(s["remaining_budget"])])
    ws.append([])
    ws.append(["Категория", "Группа", "План", "Факт", "Остаток", "Правило"])
    for line in s["lines"]:
        ws.append([line["category"], line["group"], float(line["planned"]), float(line["spent"]), float(line["remaining"]), line["carry_rule"]])

    tx_ws = wb.create_sheet("Операции")
    tx_ws.append(["Дата", "Тип", "Кто", "Категория", "Сумма", "Комментарий"])
    txs = await session.scalars(select(Transaction).options(selectinload(Transaction.category)).order_by(Transaction.tx_date.desc(), Transaction.id.desc()).limit(3000))
    for tx in txs:
        tx_ws.append([tx.tx_date.isoformat(), tx.tx_type, tx.person, tx.category.name if tx.category else "", float(tx.amount), tx.comment])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in sheet.columns:
            width = min(42, max(10, max(len(str(c.value or "")) for c in col) + 2))
            sheet.column_dimensions[get_column_letter(col[0].column)].width = width
    ws["A1"].font = Font(bold=True, size=16)
    ws["A9"].fill = PatternFill("solid", fgColor="D9EAD3")
    for c in ws[9]:
        c.font = Font(bold=True)

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    path = Path(tmp.name)
    wb.save(path)
    return path
